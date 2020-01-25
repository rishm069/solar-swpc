import os
import datetime
import sys
import collections
from openpyxl import Workbook 
from openpyxl import load_workbook

#Directories here are set for sourse data

#For Linux:
#directory = '/home/rinat/SRS'
directory = '/home/rinat/TEST'

#For Windows:
#directory = r'C:\Users\User\Desktop\handl\SRS'
#directory = r'C:\Users\User\Desktop\handl\TEST'

#Date should be un-hardcoded here
date = "20081231"
date_1 = datetime.datetime.strptime(date, "%Y%m%d")


# default() prints results in the console and can save them in .txt (might be useful for grep)
def default():
    print('stuff')
    date = "20081231"
    date_1 = datetime.datetime.strptime(date, "%Y%m%d")
    for filename in os.listdir(directory):
        fh = open(filename).read().splitlines()
        for i, line in enumerate(fh): 
            if "Nmbr Location  Lo  Area  Z   LL   NN Mag Type" in line:
            
                #Uncomment to save output to a file
            
                #sys.stdout = open('results.txt', 'a')
                #sys.stdout = open('results_data_only.txt', 'a')
            
                sh = fh[i+1:i+40]
                s = "\n".join(sh)
                sl = s.rfind('IA.')
                
                
                date_1 = date_1 + datetime.timedelta(days=1)
                end_date = date_1.strftime('%Y-%m-%d')

                #Only-results mode if uncommend if-statement below and comment it from 39 to 42
            
                #if s.startswith("None"):
                    #break
            
                print('- - - - - - - - - - - - - - - - - - - - - - -')
                print('File name: ' + str(filename))
                print('Date: ' + end_date)
                print('Data: ' + '\n')
            
                #Default: results + empty data
            
                if s.startswith("None"):
                    print('Empty dataset')
                    print('\n')
                    break
            
                
                print(line)
                print(s[:sl])

                nl = s[:sl]
                nls = list(nl.split('\n'))
                nls.pop()
                
                # for loop belov extracts values from table to a list 
                # may be used to write it into a dictionary along with 'line'
                
                for value in nls:
                    res = value.split()
                    print(line)
                    print(res) 



                #print(nls)

# sepdata() creates data.xlsx file and writes data in. In doesn't create headers
def sepdata():
    colnumber = 2
    wb = Workbook()
    wb.save('data.xlsx')
    #Date should be un-hardcoded here
    date = "20081231"
    date_1 = datetime.datetime.strptime(date, "%Y%m%d")
    for filename in os.listdir(directory):
        fh = open(filename).read().splitlines()
        for i, line in enumerate(fh): 
            if "Nmbr Location  Lo  Area  Z   LL   NN Mag Type" in line:
            
                sh = fh[i+1:i+40]
                s = "\n".join(sh)
                sl = s.rfind('IA.')
                
                date_1 = date_1 + datetime.timedelta(days=1)
                end_date = date_1.strftime('%Y-%m-%d')

                colA = "A" + str(colnumber)
                colB = "B" + str(colnumber)
                colC = "C" + str(colnumber)
                colD = "D" + str(colnumber)
                colE = "E" + str(colnumber)
                colF = "F" + str(colnumber)
                colG = "G" + str(colnumber)
                colH = "H" + str(colnumber)
                colI = "I" + str(colnumber)
                
                #Date should be un-hardcoded here
                print('Processing date: ' + end_date + '/2018-12-31')

                if s.startswith("None"):
                    break
                nl = s[:sl]
                nls = list(nl.split('\n'))
                nls.pop()
                
                #For some reason this loop (it is actually the main thing here) doesn't work in Linux environment
                for value in nls:
                    res = value.split()

                    wb = load_workbook('data.xlsx')
                    ws = wb.active
                    ws[colA] = end_date
                    wb.save('data.xlsx')

                    wb = load_workbook('data.xlsx')
                    ws = wb.active
                   
                    ws[colB] = int(res[0])
                    ws[colC] = res[1]
                    ws[colD] = int(res[2])
                    ws[colE] = int(res[3])
                    ws[colF] = res[4]
                    ws[colG] = int(res[5])                 
                    try:
                        ws[colH] = int(res[6])
                    except:
                        ws[colH] = res[6]
                    try:
                        ws[colI] = res[7]
                    except:
                        ws[colI] = ' '

                    wb.save('data.xlsx')
                    colnumber = colnumber + 1
                    colA = "A" + str(colnumber)
                    colB = "B" + str(colnumber)
                    colC = "C" + str(colnumber)
                    colD = "D" + str(colnumber)
                    colE = "E" + str(colnumber)
                    colF = "F" + str(colnumber)
                    colG = "G" + str(colnumber)
                    colH = "H" + str(colnumber)
                    colI = "I" + str(colnumber)             
              
sepdata() 
















        
   
